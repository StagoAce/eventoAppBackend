from django.shortcuts import render
from .models import user_collection, evento_collection
from django.http import HttpResponse, JsonResponse
from datetime import date, datetime
import json
from rest_framework import status
from django.views.decorators.csrf import csrf_exempt
from bson import ObjectId
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO

# Create your views here.

def index(request):
    return HttpResponse("<h1> App is running.. </h1>")

@csrf_exempt
def validate_user(request):
    if request.method == "POST":
        try:
            body = json.loads(request.body)
            if "cedula" not in body or "email" not in body:
                return JsonResponse({"error": "Cedula and email are required"}, status=400)

            user = user_collection.find_one({
                "cedula": body.get("cedula"),
                "email": body.get("email")
            })
        
            if user is not None:
                record = {
                    "cedula": user['cedula'],
                    "nombre": user['nombre'],
                    "email": user['email'],
                }
                
                return JsonResponse({"user": record}, status=200)
            else:
                return JsonResponse({"message": "This person is a new user"}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def add_user(request):
    if request.method == "POST":
        try:
            body = json.loads(request.body)
            user = user_collection.find_one({"cedula": body.get("cedula")})
        
            if user is None:
                
                try:
                    cedula = int(body.get("cedula"))
                except (ValueError, TypeError):
                    return JsonResponse({"error": "Cedula solo puede contener numeros enteros"}, status=400)
                
                records = {
                    "cedula": cedula,
                    "nombre": body.get("nombre"),
                    "apellidos": body.get("apellidos"),
                    "email": body.get("email"),
                    "rol": [body.get("rol", "usuario").lower()],
                    "eventos_out": [],
                    "eventos_in": []
                }
                
                user_collection.insert_one(records)
                return JsonResponse({"message": "New person is added"}, status=201)
            else:
                return JsonResponse({"message": "This person is not a new user"}, status=409)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

def get_all_users(request):
    try:
        users = user_collection.find()
        user_list = list(users)
        
        for user in user_list:
            if '_id' in user and isinstance(user['_id'], ObjectId):
                user['_id'] = str(user['_id'])
            if 'eventos_in' in user and isinstance(user['eventos_in'], list):
                user['eventos_in'] = [str(event_id) if isinstance(event_id, ObjectId) else event_id for event_id in user['eventos_in']]
            if 'eventos_out' in user and isinstance(user['eventos_out'], list):
                user['eventos_out'] = [str(event_id) if isinstance(event_id, ObjectId) else event_id for event_id in user['eventos_out']]
        
        users = json.dumps(user_list)
        
        return JsonResponse(user_list, safe=False, status=status.HTTP_200_OK) 
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def get_user(request, user_cedula):
    try:
        user = user_collection.find_one({"cedula": user_cedula})
        
        if user is not None:
            if '_id' in user and isinstance(user['_id'], ObjectId):
                user['_id'] = str(user['_id'])
            if 'eventos_in' in user and isinstance(user['eventos_in'], list):
                user['eventos_in'] = [str(event_id) if isinstance(event_id, ObjectId) else event_id for event_id in user['eventos_in']]
            if 'eventos_out' in user and isinstance(user['eventos_out'], list):
                user['eventos_out'] = [str(event_id) if isinstance(event_id, ObjectId) else event_id for event_id in user['eventos_out']]
            return JsonResponse(user, status=200)
        else:
            return JsonResponse({"error": "User not found"}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@csrf_exempt
def subscribe_evento(request, user_cedula):
    if request.method == "POST":
        try:
            user = user_collection.find_one({"cedula": user_cedula})
            if user is not None:
                body = json.loads(request.body)
                id_evento = body.get("id_evento")
                
                evento = evento_collection.find_one({"_id": ObjectId(id_evento)})
                if evento is not None:
                    #aCTUALIZAR INFORMACION DE EVENTOS DEL USUARIO
                    filtro = {"_id": user["_id"]}
                    actualizacion = {"$push": {"eventos_out": evento["_id"]}}
                    user_collection.update_one(filtro, actualizacion)
                    
                    #ACTUALIZAR INFORMACION DE ASISTENTES DEL EVENTO
                    filtro = {"_id": evento["_id"]}
                    actualizacion = {"$push": {"asistentes": user["_id"]}}
                    evento_collection.update_one(filtro, actualizacion)
                    
                    return JsonResponse({"message": "Subscripcion exitosa"}, status=200)
                else:
                    return JsonResponse({"error": "Evento not found"}, status=404)
            else:
                return JsonResponse({"error": "User not found"}, status=404)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

from bson import ObjectId  # Asegúrate de importar ObjectId
from django.http import JsonResponse
import json

@csrf_exempt        
def unsubscribe_evento(request, user_cedula):
    if request.method == "POST":
        try:
            user = user_collection.find_one({"cedula": user_cedula})
            if user is not None:
                body = json.loads(request.body)
                id_evento = ObjectId(body.get("id_evento"))
                
                evento = evento_collection.find_one({"_id": id_evento})
                if evento is not None:
                    if user["_id"] in evento["asistentes"]:
                        evento_collection.update_one(
                            {"_id": id_evento},  # Usar el ObjectId directamente
                            {"$pull": {"asistentes": user["_id"]}}  # ID del asistente a remover
                        )
                        
                        # Eliminar el evento de la lista de eventos del usuario
                        user_collection.update_one(
                            {"_id": user["_id"]},  # ID del usuario
                            {"$pull": {"eventos_out": id_evento}}  # Usar el ObjectId
                        )
                        
                        return JsonResponse({"message": "Se desuscribió de manera exitosa"}, status=200)
                    else:
                        return JsonResponse({"error": "El usuario no se encuentra en este evento"}, status=404)
                else:
                    return JsonResponse({"error": "Evento no encontrado"}, status=404)
            else:
                return JsonResponse({"error": "Usuario no encontrado"}, status=404)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

def get_eventos_user(request, user_cedula):
    try:
        user = user_collection.find_one({"cedula": user_cedula})
        
        if user is not None:
            # Obtener la lista de eventos a los que el usuario está inscrito
            eventos_out = user.get("eventos_out", [])
            # Convertir cada ID de la lista a ObjectId para la consulta
            evento_ids = [ObjectId(event_id) for event_id in eventos_out]

            # Buscar los eventos en la colección de eventos
            eventos = list(evento_collection.find({"_id": {"$in": evento_ids}}))

            # Serializar cada evento y eliminar la lista de asistentes
            eventos_serializados = [
            {key: value for key, value in serialize_evento(evento).items() if key != 'asistentes'}
            for evento in eventos]

            return JsonResponse(eventos_serializados, safe=False, status=200)  # Devuelve la lista de eventos serializados
        else:
            return JsonResponse({"error": "Usuario no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
#------------------------ AREA DE EVENTOS -----------------------------------
@csrf_exempt
def create_evento(request, cedula_cliente):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            user = user_collection.find_one({"cedula": cedula_cliente})
            
            if user is not None:
                fecha_inicio_str = body.get("fecha_inicio")
                fecha_finalizacion_str = body.get("fecha_finalizacion")
                try:
                    fecha_inicio = datetime.fromisoformat(fecha_inicio_str)
                    fecha_finalizacion = datetime.fromisoformat(fecha_finalizacion_str)
                    
                    if fecha_inicio.date() > date.today():
                        estado = "activo"
                    elif fecha_inicio.date() <= date.today() <= fecha_finalizacion.date():
                        estado = "en curso"
                    else:
                        estado = "finalizado"

                    existing_event = evento_collection.find_one({
                        "nombre": body.get("nombre"),
                        "organizador": user.get("nombre", "") + " " + user.get("apellidos", ""),
                        "lugar": body.get("lugar"),
                        "fecha_inicio": fecha_inicio,
                        "fecha_finalizacion": fecha_finalizacion
                    })
                    
                    if existing_event:
                        return JsonResponse({"message": "El evento ya existe."}, status=409)
                
                    records = {
                        "nombre": body.get("nombre"),
                        "organizador": user.get("nombre", "") + " " + user.get("apellidos", ""),
                        "lugar": body.get("lugar"),
                        "direccion": body.get("direccion"),
                        "fecha_inicio": fecha_inicio,
                        "fecha_finalizacion": fecha_finalizacion,
                        "descripcion": body.get("descripcion"),
                        "asistentes": [],
                        "estado": estado
                    }
                    
                    result = evento_collection.insert_one(records)
                                
                    filtro = {"_id": user["_id"]}
                    actualizacion = {"$push": {"eventos_in": result.inserted_id}}
                    user_collection.update_one(filtro, actualizacion)
                    
                    return JsonResponse({"message": "Creacion exitosa"}, status=201)
                    
                except ValueError as ve:
                    return JsonResponse({"message": "Error con las fechas: " + str(ve)}, status=400)
            else:
                return JsonResponse({"message": "Person doesn't exist"}, status=409)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    else:
        return JsonResponse({"error": "Metodo incorrecto"}, status=405)
    
#Función para serializar un evento, convirtiendo ObjectId y datetime a tipos JSON serializables
def serialize_evento(evento):
    evento['_id'] = str(evento['_id'])
    if 'fecha_inicio' in evento:
        evento['fecha_inicio'] = evento['fecha_inicio'].isoformat()
    if 'fecha_finalizacion' in evento:
        evento['fecha_finalizacion'] = evento['fecha_finalizacion'].isoformat()
    if 'asistentes' in evento and isinstance(evento['asistentes'], list):
        evento['asistentes'] = [str(asistente) for asistente in evento['asistentes']]
    
    return evento

def get_all_eventos(request):
    try:
        eventos = evento_collection.find()
        eventos_list = list(eventos)
        # Serializar cada evento
        eventos_list = [serialize_evento(evento) for evento in eventos_list]
        
        return JsonResponse(eventos_list, safe=False, status=status.HTTP_200_OK) 
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def get_all_eventos_by_client(request, user_cedula):
    try:
        user = user_collection.find_one({"cedula": user_cedula})
        if user is not None:
            eventos_in = user.get("eventos_in", [])
            
            evento_ids = [ObjectId(event_id) for event_id in eventos_in]

            # Buscar los eventos en la colección de eventos
            eventos = list(evento_collection.find({"_id": {"$in": evento_ids}}))
            
            # Serializar cada evento, eliminar la lista de asistentes y agregar la cantidad de asistentes
            eventos_serializados = [
                {**{key: value for key, value in serialize_evento(evento).items() if key != 'asistentes'},
                 'num_asistentes': len(evento.get("asistentes", []))}  # Añadir número de asistentes
                for evento in eventos
            ]

            return JsonResponse(eventos_serializados, safe=False, status=200)
        else:
            return JsonResponse({"error": "Usuario no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def get_evento(request, evento_id):
    try:
        evento = evento_collection.find_one({"_id": ObjectId(evento_id)})
        
        if evento is not None:
            evento = serialize_evento(evento)
            return JsonResponse(evento, status=200)
        else:
            return JsonResponse({"error": "Evento no encontrado"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
                
def generar_reporte(request):
    eventos = evento_collection.find()
    print("Hola  1",eventos)
    root = ET.Element("eventos")

    for evento in eventos:
        evento_elem = ET.SubElement(root, "evento")

        # Añadir los campos del evento
        ET.SubElement(evento_elem, "nombre").text = evento.get("nombre", "Sin Nombre")
        ET.SubElement(evento_elem, "organizador").text = evento.get("organizador", "Desconocido")
        ET.SubElement(evento_elem, "lugar").text = evento.get("lugar", "Desconocido")
        ET.SubElement(evento_elem, "direccion").text = evento.get("direccion", "Desconocida")
        
        # Convertir las fechas a un formato legible
        fecha_inicio = evento.get("fecha_inicio")
        fecha_finalizacion = evento.get("fecha_finalizacion")
        if fecha_inicio:
            fecha_inicio = evento['fecha_inicio']
            fecha_inicio_formateada = fecha_inicio.strftime("%Y-%m-%d %H:%M:%S")
        if fecha_finalizacion:
            fecha_finalizacion = evento['fecha_inicio']
            fecha_finalizacion_formateada = fecha_finalizacion.strftime("%Y-%m-%d %H:%M:%S")
        
        ET.SubElement(evento_elem, "fecha_inicio").text = fecha_inicio_formateada if fecha_inicio_formateada else "Fecha no disponible"
        ET.SubElement(evento_elem, "fecha_finalizacion").text = fecha_finalizacion_formateada if fecha_finalizacion_formateada else "Fecha no disponible"
        
        ET.SubElement(evento_elem, "descripcion").text = evento.get("descripcion", "Sin descripción")

    # Crear el árbol XML
    tree = ET.ElementTree(root)
  
    # Si el usuario solicita el reporte en formato Excel
    if request.GET.get('format') == 'excel':
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reporte de Eventos"
        
        headers = ["Nombre", "Organizador", "Lugar", "Dirección", "Fecha de Inicio", "Fecha de Finalización", "Descripción"]

        # Estilo de encabezado (color de fondo y negrita)
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)

        # Agregar encabezados y aplicar estilo
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font


        for evento in root.findall('evento'):
            fila = [
                evento.find("nombre").text,
                evento.find("organizador").text,
                evento.find("lugar").text,
                evento.find("direccion").text,
                evento.find("fecha_inicio").text,
                evento.find("fecha_finalizacion").text,
                evento.find("descripcion").text,
            ]
            sheet.append(fila)

        # Ajustar el tamaño de las columnas según el contenido
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Ajustar el texto para que no se desborde (ajustar el texto dentro de las celdas)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)


        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=reporte_eventos.xlsx"
        return response

    # Si el usuario solicita el reporte en formato XML
    response = HttpResponse(content_type='application/xml')
    tree.write(response, encoding="unicode", xml_declaration=True)
    return response    

        
    
